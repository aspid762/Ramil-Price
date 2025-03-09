import openpyxl
from models import db, Product, Customer, Stock
from datetime import datetime, timedelta, timedelta
from sqlalchemy.exc import SQLAlchemyError
import logging

logger = logging.getLogger('excel_import')

def import_products_from_excel(file_path):
    """
    Импортирует товары из Excel-файла.
    
    Args:
        file_path (str): Путь к Excel-файлу
    
    Returns:
        dict: Результат импорта
    """
    try:
        # Открываем Excel-файл
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Получаем заголовки столбцов
        headers = [cell.value for cell in ws[1]]
        
        # Проверяем, что файл содержит необходимые столбцы
        required_columns = ['Категория', 'Наименование', 'Характеристики', 'Цена за тонну (₽)', 'Вес п.м. (кг)']
        for column in required_columns:
            if column not in headers:
                return {
                    'success': False,
                    'error': f'В файле отсутствует столбец "{column}"',
                    'errors': [],
                    'imported': 0,
                    'updated': 0
                }
        
        # Получаем индексы столбцов
        category_idx = headers.index('Категория')
        name_idx = headers.index('Наименование')
        characteristics_idx = headers.index('Характеристики')
        price_per_ton_idx = headers.index('Цена за тонну (₽)')
        weight_per_meter_idx = headers.index('Вес п.м. (кг)')
        
        # Импортируем данные
        imported = 0
        updated = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Проверяем, что строка не пустая
                if not any(row):
                    continue
                
                # Получаем данные из строки
                category = row[category_idx]
                name = row[name_idx]
                characteristics = row[characteristics_idx]
                
                # Проверяем обязательные поля
                if not category or not name or not characteristics:
                    errors.append(f'Строка {row_idx}: Не заполнены обязательные поля')
                    continue
                
                # Преобразуем числовые значения, обрабатывая возможные ошибки
                try:
                    # Проверяем, что значения не None и не строки с заголовками
                    price_per_ton_value = row[price_per_ton_idx]
                    if price_per_ton_value is None or isinstance(price_per_ton_value, str) and price_per_ton_value.strip() == '':
                        errors.append(f'Строка {row_idx}: Не указана цена за тонну')
                        continue
                    
                    # Проверяем, что значение не является заголовком
                    if isinstance(price_per_ton_value, str) and any(keyword in price_per_ton_value.lower() for keyword in ['цена', 'тонн', '₽']):
                        errors.append(f'Строка {row_idx}: Значение "{price_per_ton_value}" не является числом')
                        continue
                    
                    price_per_ton = float(price_per_ton_value)
                    
                    weight_per_meter_value = row[weight_per_meter_idx]
                    if weight_per_meter_value is None or isinstance(weight_per_meter_value, str) and weight_per_meter_value.strip() == '':
                        errors.append(f'Строка {row_idx}: Не указан вес погонного метра')
                        continue
                    
                    # Проверяем, что значение не является заголовком
                    if isinstance(weight_per_meter_value, str) and any(keyword in weight_per_meter_value.lower() for keyword in ['вес', 'кг', 'метр']):
                        errors.append(f'Строка {row_idx}: Значение "{weight_per_meter_value}" не является числом')
                        continue
                    
                    weight_per_meter = float(weight_per_meter_value)
                except (ValueError, TypeError) as e:
                    errors.append(f'Строка {row_idx}: Ошибка преобразования числовых значений: {str(e)}')
                    continue
                
                # Проверяем, существует ли уже такой товар
                existing_product = Product.query.filter_by(
                    category=category,
                    name=name,
                    characteristics=characteristics
                ).first()
                
                if existing_product:
                    # Если товар существует, обновляем его
                    if existing_product.price_per_ton != price_per_ton:
                        # Сохраняем старую цену в историю
                        price_history = PriceHistory(
                            product_id=existing_product.id,
                            price_per_ton=existing_product.price_per_ton,
                            previous_price=None,
                            created_at=datetime.utcnow()
                        )
                        db.session.add(price_history)
                    
                    existing_product.price_per_ton = price_per_ton
                    existing_product.weight_per_meter = weight_per_meter
                    updated += 1
                else:
                    # Если товар не существует, создаем новый
                    product = Product(
                        category=category,
                        name=name,
                        characteristics=characteristics,
                        price_per_ton=price_per_ton,
                        weight_per_meter=weight_per_meter
                    )
                    db.session.add(product)
                    imported += 1
            except Exception as e:
                errors.append(f'Строка {row_idx}: {str(e)}')
        
        # Сохраняем изменения в базе данных
        db.session.commit()
        
        return {
            'success': True,
            'errors': errors,
            'imported': imported,
            'updated': updated
        }
    except Exception as e:
        logger.error(f"Error importing products: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'errors': [],
            'imported': 0,
            'updated': 0
        }

def import_customers_from_excel(file_path):
    """Импортирует заказчиков из Excel-файла."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Пропускаем заголовок
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    imported_count = 0
    updated_count = 0
    errors = []
    
    for row in rows:
        try:
            # Проверяем, что строка не пустая
            if not row[0]:
                continue
            
            name = row[0]
            phone = row[1] if row[1] else ''
            address = row[2] if row[2] else ''
            margin = float(row[3]) if row[3] else 0.0
            delivery_fee = float(row[4]) if row[4] else 0.0
            
            # Проверяем, существует ли уже такой заказчик
            existing_customer = Customer.query.filter_by(name=name).first()
            
            if existing_customer:
                # Обновляем существующего заказчика
                existing_customer.phone = phone
                existing_customer.address = address
                existing_customer.margin = margin
                existing_customer.delivery_fee = delivery_fee
                updated_count += 1
            else:
                # Создаем нового заказчика
                customer = Customer(
                    name=name,
                    phone=phone,
                    address=address,
                    margin=margin,
                    delivery_fee=delivery_fee
                )
                db.session.add(customer)
                imported_count += 1
        except Exception as e:
            errors.append(f"Ошибка в строке {rows.index(row) + 2}: {str(e)}")
    
    try:
        db.session.commit()
        return {
            'success': True,
            'imported': imported_count,
            'updated': updated_count,
            'errors': errors
        }
    except SQLAlchemyError as e:
        db.session.rollback()
        return {
            'success': False,
            'error': str(e),
            'errors': errors
        }

def import_stock_from_excel(file_path):
    """Импортирует товары на склад из Excel-файла."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Пропускаем заголовок
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    imported_count = 0
    updated_count = 0
    errors = []
    
    for row in rows:
        try:
            # Проверяем, что строка не пустая
            if not row[0]:
                continue
            
            category = row[0]
            name = row[1]
            characteristics = row[2]
            quantity = float(row[3])
            purchase_price = float(row[4])
            
            # Создаем новую позицию на складе
            stock = Stock(
                category=category,
                name=name,
                characteristics=characteristics,
                quantity=quantity,
                reserved_quantity=0.0,
                purchase_price=purchase_price,
                received_at=datetime.now()
            )
            db.session.add(stock)
            imported_count += 1
            
            # Добавляем запись о движении товара
            from models import StockMovement
            movement = StockMovement(
                stock_id=stock.id,
                operation_type='поступление',
                quantity=quantity,
                purchase_price=purchase_price,
                created_at=datetime.now()
            )
            db.session.add(movement)
        except Exception as e:
            errors.append(f"Ошибка в строке {rows.index(row) + 2}: {str(e)}")
    
    try:
        db.session.commit()
        return {
            'success': True,
            'imported': imported_count,
            'updated': updated_count,
            'errors': errors
        }
    except SQLAlchemyError as e:
        db.session.rollback()
        return {
            'success': False,
            'error': str(e),
            'errors': errors
        } 