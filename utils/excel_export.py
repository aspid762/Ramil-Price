import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

def create_excel_report(title, headers, data, filename=None):
    """
    Создает Excel-отчет с указанными данными.
    
    Args:
        title (str): Заголовок отчета
        headers (list): Список заголовков столбцов
        data (list): Список строк с данными (каждая строка - список значений)
        filename (str, optional): Имя файла. Если не указано, генерируется автоматически.
    
    Returns:
        BytesIO: Объект с данными Excel-файла
    """
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет"
    
    # Добавляем заголовок отчета
    ws.merge_cells('A1:{}1'.format(get_column_letter(len(headers))))
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Добавляем дату формирования отчета
    ws.merge_cells('A2:{}2'.format(get_column_letter(len(headers))))
    ws['A2'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Добавляем заголовки столбцов
    header_row = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Добавляем данные
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left')
    
    # Автоматически регулируем ширину столбцов
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def export_products_to_excel(products):
    """Экспортирует список товаров в Excel."""
    headers = ['Категория', 'Наименование', 'Характеристики', 'Цена за тонну (₽)', 'Вес п.м. (кг)', 'Цена за п.м. (₽)']
    
    data = [
        [
            product.category,
            product.name,
            product.characteristics,
            product.price_per_ton,
            product.weight_per_meter,
            product.price_per_meter
        ]
        for product in products
    ]
    
    return create_excel_report(
        title="Прайс-лист металлопроката",
        headers=headers,
        data=data,
        filename="price_list.xlsx"
    )

def export_customers_to_excel(customers):
    """Экспортирует список заказчиков в Excel."""
    headers = ['Наименование', 'Телефон', 'Адрес', 'Наценка (%)', 'Стоимость доставки (₽)']
    
    data = [
        [
            customer.name,
            customer.phone,
            customer.address,
            customer.margin,
            customer.delivery_fee
        ]
        for customer in customers
    ]
    
    return create_excel_report(
        title="Список заказчиков",
        headers=headers,
        data=data,
        filename="customers.xlsx"
    )

def export_orders_to_excel(orders):
    """Экспортирует список заказов в Excel."""
    headers = ['ID', 'Заказчик', 'Дата создания', 'Статус', 'Ожидаемая отгрузка', 'Сумма (₽)', 'Примечания']
    
    data = [
        [
            order.id,
            order.customer.name,
            order.created_at.strftime('%d.%m.%Y'),
            order.status,
            order.expected_shipping.strftime('%d.%m.%Y') if order.expected_shipping else '',
            order.total_cost,
            order.notes or ''
        ]
        for order in orders
    ]
    
    return create_excel_report(
        title="Список заказов",
        headers=headers,
        data=data,
        filename="orders.xlsx"
    )

def export_stock_to_excel(stock_items):
    """Экспортирует список товаров на складе в Excel."""
    headers = ['Категория', 'Наименование', 'Характеристики', 'Количество', 'Зарезервировано', 'Доступно', 'Закупочная цена (₽)', 'Дата поступления']
    
    data = [
        [
            item.category,
            item.name,
            item.characteristics,
            item.quantity,
            item.reserved_quantity,
            item.available_quantity,
            item.purchase_price,
            item.received_at.strftime('%d.%m.%Y')
        ]
        for item in stock_items
    ]
    
    return create_excel_report(
        title="Складской учет",
        headers=headers,
        data=data,
        filename="stock.xlsx"
    )

def export_order_details_to_excel(order):
    """Экспортирует детали заказа в Excel."""
    # Информация о заказе
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заказ №{order.id}"
    
    # Заголовок
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Заказ №{order.id} от {order.created_at.strftime('%d.%m.%Y')}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Информация о заказчике
    ws['A3'] = "Заказчик:"
    ws['B3'] = order.customer.name
    ws['A4'] = "Телефон:"
    ws['B4'] = order.customer.phone
    ws['A5'] = "Адрес:"
    ws['B5'] = order.customer.address
    
    # Информация о заказе
    ws['D3'] = "Статус:"
    ws['E3'] = order.status
    ws['D4'] = "Ожидаемая отгрузка:"
    ws['E4'] = order.expected_shipping.strftime('%d.%m.%Y') if order.expected_shipping else 'Не указана'
    ws['D5'] = "Примечания:"
    ws['E5'] = order.notes or ''
    
    # Заголовки таблицы позиций
    headers = ['№', 'Наименование', 'Характеристики', 'Количество', 'Цена (₽)', 'Сумма (₽)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Данные о позициях заказа
    for idx, item in enumerate(order.items, 1):
        row_idx = 7 + idx
        
        # Определяем наименование и характеристики
        if item.product:
            name = f"{item.product.category} {item.product.name}"
            characteristics = item.product.characteristics
        elif item.stock:
            name = f"{item.stock.category} {item.stock.name}"
            characteristics = item.stock.characteristics
        else:
            name = "Неизвестный товар"
            characteristics = ""
        
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=characteristics)
        ws.cell(row=row_idx, column=4, value=item.quantity)
        ws.cell(row=row_idx, column=5, value=item.selling_price)
        ws.cell(row=row_idx, column=6, value=item.quantity * item.selling_price)
    
    # Итоговая сумма
    total_row = 7 + len(order.items) + 2
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws.cell(row=total_row, column=1, value="ИТОГО:")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=order.total_cost)
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    
    # Доставка
    if order.customer.delivery_fee > 0:
        delivery_row = total_row - 1
        ws.merge_cells(f'A{delivery_row}:E{delivery_row}')
        ws.cell(row=delivery_row, column=1, value="Доставка:")
        ws.cell(row=delivery_row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=delivery_row, column=6, value=order.customer.delivery_fee)
    
    # Автоматически регулируем ширину столбцов
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def export_stock_movements_to_excel(movements, start_date=None, end_date=None):
    """
    Экспортирует движения товаров на складе в Excel.
    
    Args:
        movements (list): Список объектов StockMovement
        start_date (datetime, optional): Начальная дата периода
        end_date (datetime, optional): Конечная дата периода
    
    Returns:
        BytesIO: Объект с данными Excel-файла
    """
    # Создаем заголовок отчета
    title = "Отчет о движении товаров на складе"
    if start_date and end_date:
        title += f" за период с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
    
    # Определяем заголовки столбцов
    headers = [
        "№ п/п",
        "Дата",
        "Категория",
        "Наименование",
        "Характеристики",
        "Тип операции",
        "Количество",
        "Цена, ₽",
        "Сумма, ₽",
        "Заказ №"
    ]
    
    # Подготавливаем данные
    data = []
    for idx, movement in enumerate(movements, 1):
        # Получаем информацию о товаре
        stock = movement.stock
        
        # Формируем строку данных
        row = [
            idx,
            movement.created_at.strftime('%d.%m.%Y %H:%M'),
            stock.category if stock else "",
            stock.name if stock else "",
            stock.characteristics if stock else "",
            movement.operation_type,
            movement.quantity,
            movement.purchase_price,
            movement.quantity * movement.purchase_price,
            movement.order_id if movement.order_id else ""
        ]
        
        data.append(row)
    
    # Создаем отчет
    return create_excel_report(title, headers, data)

def export_customers_to_excel(customers):
    """
    Экспортирует список заказчиков в Excel.
    
    Args:
        customers (list): Список объектов Customer
    
    Returns:
        BytesIO: Объект с данными Excel-файла
    """
    # Создаем заголовок отчета
    title = "Список заказчиков"
    
    # Определяем заголовки столбцов
    headers = [
        "№ п/п",
        "Наименование",
        "Телефон",
        "Адрес",
        "Наценка, %",
        "Стоимость доставки, ₽"
    ]
    
    # Подготавливаем данные
    data = []
    for idx, customer in enumerate(customers, 1):
        row = [
            idx,
            customer.name,
            customer.phone,
            customer.address,
            customer.margin,
            customer.delivery_fee
        ]
        
        data.append(row)
    
    # Создаем отчет
    return create_excel_report(title, headers, data)

def export_orders_to_excel(orders):
    """
    Экспортирует список заказов в Excel.
    
    Args:
        orders (list): Список объектов Order
    
    Returns:
        BytesIO: Объект с данными Excel-файла
    """
    # Создаем заголовок отчета
    title = "Список заказов"
    
    # Определяем заголовки столбцов
    headers = [
        "№ заказа",
        "Дата создания",
        "Заказчик",
        "Статус",
        "Сумма, ₽",
        "Ожидаемая отгрузка",
        "Фактическая отгрузка"
    ]
    
    # Подготавливаем данные
    data = []
    for order in orders:
        row = [
            order.id,
            order.created_at.strftime('%d.%m.%Y'),
            order.customer.name if order.customer else "",
            order.status,
            order.total_cost,
            order.expected_shipping.strftime('%d.%m.%Y') if order.expected_shipping else "",
            order.actual_shipping.strftime('%d.%m.%Y') if order.actual_shipping else ""
        ]
        
        data.append(row)
    
    # Создаем отчет
    return create_excel_report(title, headers, data) 